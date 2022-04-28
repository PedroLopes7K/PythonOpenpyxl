#LER OS DADOS DE UMA PLANILHA TODA

import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('Desafio2.xlsx')
sheet = wb['Sheet1']
linhas = sheet.max_row
colunas = sheet.max_column
#print(linhas)
#print(colunas)
a = sheet['A1']
b = sheet['B1']
c = sheet['C1']
d = sheet['D1']

for row in range(2, linhas + 1):
    print('==== Linha %s ==== ' % (row) )
    print(a.value,':')
    print(sheet['A%s' % (row)].value)
    print('------')
    print(b.value,':')
    print(sheet['B%s' % (row)].value)
    print('------')
    print(c.value,':')
    print(sheet['C%s' % (row)].value)
    print('------')
    print(d.value,':')
    print(sheet['D%s' % (row)].value)
    print('==========')
    print('')
    
''' print('==== Linha %s ==== ' % (row) )
    for colun in range(1, colunas + 1):
        c = get_column_letter(colun)
        print(sheet['%s%s' % ( c, row)].value)
    print('==========')
    print('')''' 
