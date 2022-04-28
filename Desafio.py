import openpyxl
wb = openpyxl.load_workbook('mudou.xlsx')
sheet = wb['Sheet1']
linhas = sheet.max_row
colunas = sheet.max_column
  
print(linhas)
print(colunas)
for i in range(1,linhas):
    print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value )
    if sheet.cell(row=i, column=1).value == 'Camisa Amarela':
        sheet.cell(row=i, column=2).value = 30
    if sheet.cell(row=i, column=1).value == 'Camisa Azul':
        sheet.cell(row=i, column=2).value = 45
print('Sucesso!')
wb.save('Desafio.xlsx')
