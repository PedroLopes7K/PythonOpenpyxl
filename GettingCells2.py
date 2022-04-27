import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1'] # Get a sheet from the workbook.
#sheet.cell(row=1, column=2)

#sheet.cell(row=1, column=2).value

for i in range(1, 8): # Go through every other row:
    if sheet.cell(row=1, column=i).value != None:
        print(i, sheet.cell(row=1, column=i).value)
    if sheet.cell(row=1, column=i).value == 'red':
        sheet.cell(row=1, column=i).value = 'Mudou'
        print('Achou')
    else:
      continue
wb.save('teste.xlsx')
